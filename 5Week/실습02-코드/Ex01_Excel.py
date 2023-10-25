"""
###############################################################################
# Openpyxl vs. MS COM 모듈
###############################################################################
import os
import time
import openpyxl
import win32com.client      # pip install pywin32


def load_xls_by_openpyxl(fname):
    openpyxl.load_workbook(fname)


def load_xls_by_mscom(fname):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.Workbooks.Open(fname)
    excel.Quit()


if __name__ == '__main__':
    fname = os.path.join(os.getcwd(), 'bom_whole.xlsx')

    # OpenPyXl 불러오기 시간
    t0 = time.time()
    for i in range(5):
        load_xls_by_openpyxl(fname)  # 15.1초/회 소요
    print('OpenPyXl 불러오기 : {:.1f} 초'.format((time.time()-t0)/5))

    # Win2Com 불러오기 시간
    t0 = time.time()
    for i in range(5):
        load_xls_by_mscom(fname)       # 5.2초/회 소요, full path filename
    print('Win2Com 불러오기 : {:.1f} 초'.format((time.time()-t0)/5))


###############################################################################
# Openpyxl 모듈(1) - 기존문서를 불러와서 작업하고 저장하기
###############################################################################
import openpyxl

xls = openpyxl.load_workbook('재고장.xlsx')     # Workbook 열기

xls.sheetnames       # 시트목록을 리스트로 반환 ['물류재고장', '입출고내역', '랙번호']
sheet = xls.active   # 현재 활성화된 sheet를 가져옴
sheet.title          # '물류재고장'
sheet = xls['랙번호'] # '랙번호' sheet 가져오기
sheet.max_row        # 503
sheet.max_column     # 1

sheet['A2']             # <Cell '랙번호'.A2>
sheet['A2'].value       # '0-1-1'
sheet['A2'].coordinate  # 'A2'

sheet.cell(row=2, column=1)     # 숫자 인덱스를 통해 루프를 돌릴 수 있다
sheet.cell(2, 1).value
sheet.cell(2, 1).coordinate
for i in range(1, 8, 2):
    print(i, sheet.cell(i, 1).value)
    # 1 랙번호
    # 3 0-1-2
    # 5 0-2-1
    # 7 0-2-3


# 셀값 가져오기
sheet = xls['물류재고장']
area = sheet['A1:C3']
# ((<Cell '물류재고장'.A1>, <Cell '물류재고장'.B1>, <Cell '물류재고장'.C1>),
#  (<Cell '물류재고장'.A2>, <Cell '물류재고장'.B2>, <Cell '물류재고장'.C2>),
#  (<Cell '물류재고장'.A3>, <Cell '물류재고장'.B3>, <Cell '물류재고장'.C3>))
area[0][1]              # <Cell '물류재고장'.B1>
for row in area:
    for cell in row:
        print(cell.value, end='\t\t')
    print()

row = sheet[1]
str_row = []
for cell in row:
    str_row.append(cell.value)      # Print, print -> 줄바꿈 (end='\n'), print(, ) -> 공백 (sep=',')
print(", ".join(str_row))

rows = sheet[2:5]
for row in rows:
    for cell in row:
        print(cell.value, end=' / ')
    print()

col = sheet['A']
for cell in col:
    print(cell.value)

cols = sheet['A:B']
trans_cols = zip(*cols)     # zip + unpacking --> 행과 열의 transpose
for row in trans_cols:
    for cell in row:
        print(cell.value, end='\t')
    print()

xls.close()

# 리스트의 행과 열을 바꾸기 : zip + unpacking
A = [[1, 2, 3], [4, 5, 6]]
list(zip(*A))           # [(1, 4), (2, 5), (3, 6)]

###############################################################################
# Openpyxl 모듈(2) - 새 문서 만들고 저장하기
###############################################################################
from openpyxl import Workbook

book = Workbook()                    # 새로운 엑셀파일 생성

book.sheetnames                      # default로 'Sheet' 시트 생성
sheet = book.active                  # 현재 활성시트
sheet.title = 'Stock'               # sheetname 변경

book.create_sheet('가나다')           # 마지막 위치에 '가나타' 시트 추가
book.create_sheet('시트2', 1)         # 해당위치 뒤에 '시트2' 시트 추가

sheet['A1'].value = 'Rack'
# sheet['A1'] = 'Rack'             # value 생략 가능
sheet.cell(2, 1).value = '0-0-1'    # 인덱스가 (1, 1)부터 가능, value 생략불가

sheet.merge_cells('B1:C1')          # 셀 합치기
sheet['B1'].value = 'B1과 C1 결합'
sheet.merge_cells('A4:C6')
sheet['A4'].value = 'A4:C6 범위 합치기'

sheet['D2'].value = 2
sheet['D4'].value = 4
sheet['D6'].value = '=SUM(D1:D5)'

book.save('재고장_new.xlsx')          # 저장
book.close()                         # 닫기

"""


###############################################################################
# [실습3] Openpyxl 모듈(3) - 셀서식 지정하기
###############################################################################
# import openpyxl
from openpyxl import *
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# 상수
TITLE_CELL_COLOR = "AA8866"

wb = openpyxl.load_workbook(r'C:\Users\user\SmartFactory\5Week\실습02-코드\재고장.xlsx')
ws_raw = wb['물류재고장']
# ws.append(['1-2-3', 'SI1234', '추가된 사출성형 품목', '사출', 6000, '2021-03-29'])

ws = wb.copy_worksheet(ws_raw)
ws.title = '물류재고장_style'

# 틀 고정
ws.freeze_panes = "C2"

# 열 크기 지정
col_widths = {'A': 8, 'B': 13, 'C': 50, 'D': 10, 'E': 10, 'F': 10, 'G': 15}
for pos, width in col_widths.items():
    ws.column_dimensions[pos].width = width

# 행 높이 지정
for i in range(1, ws.max_row+1):
    ws.row_dimensions[i].height = 20
    ws.cell(i, 5).number_format = '#,##0'
    if i != 1 and int(ws.cell(i, 5).value) < 1000:
        ws.cell(i, 5).font = Font(bold=True)

# 폰트 지정
font_header = Font(name='맑은 고딕', size=12, bold=True, color='FFFFFF')
for row in ws['A1':'G1']:
    for cell in row:
        cell.fill = PatternFill(patternType='solid', fgColor=TITLE_CELL_COLOR)
        cell.alignment = Alignment(horizontal='distributed')
        cell.font = font_header

# 셀 테두리 지정
side = Side(style='thin', color='000000')
border = Border(left=side, right=side, top=side, bottom=side)
for row in ws:
    for cell in row:
        cell.border = border

# 행과 열 숨기기
# ws.column_dimensions['A'].hidden = True     # 열 숨기기
# # ws.column_dimensions['A'].hidden = False    # 숨긴 열 표시하기
# ws.row_dimensions[1].hidden = True

wb.save('재고장_style.xlsx')
wb.close()

"""
###############################################################################
# Openpyxl 모듈(4) - 간단한 chart 그리기
###############################################################################
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
import random


wb = Workbook()
ws = wb.active

for _ in range(1, 6):
    num = random.randint(1, 6)
    ws.append([num])      # A열에 0~9 입력

values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)     # 열 우선에 주의
chart = BarChart()      # LineChart()
chart.add_data(values)
chart.title = '차트 제목';  chart.x_axis.title = 'x축 제목'; chart.y_axis.title = 'y축 제목'
ws.add_chart(chart, "C1")

wb.save("재고장_SimpleChart.xlsx")
wb.close()


###############################################################################
# [실습4] Openpyxl 모듈(5) - 다양한 chart 그리기
###############################################################################
import openpyxl
from openpyxl.chart import Reference, BarChart, LineChart, AreaChart, PieChart, RadarChart


wb = openpyxl.load_workbook('재고장.xlsx')
ws = wb['물류재고장']

# chart 데이터 참조범위 지정
data1 = Reference(ws, min_col=5, min_row=1, max_col=6, max_row=12)  # 재고량(5열), 주문량(6열)
# data2 = Reference(ws, 5, 2, 5, 12)
labels = Reference(ws, 2, 2, 2, 12)                                 # 부품코드(2열)

# 차트 종류 지정
# chart = BarChart()    # 막대그래프
# chart.type = 'bar'    # 가로(bar), 세로(col) - 디폴트값
# chart = LineChart()     # 꺾은 선 그래프
# chart = AreaChart()     # 영역형 그래프
# chart = PieChart()      # 원형 차트
chart = RadarChart()    # 방사형 차트

# chart.grouping = 'stacked'  # 누적
# chart.overlap = 100       # for barchart: 100

chart.add_data(data1, titles_from_data=True)    # title...=True: 첫 셀값을 계열값(범례)로 사용
# chart.add_data(data2)
chart.set_categories(labels)

chart.title = '재고현황'              # 차트 제목
# chart.x_axis.title = '랙번호'        # x축 제목
# chart.y_axis.title = '재고량'        # y축 제목

chart.height = 10                   # 차트 가로 크기
chart.width = 15                    # 차트 세로 크기

ws.add_chart(chart, 'B14')          # chart를 시트의 B14에 삽입

wb.save('재고장_radarchart.xlsx')
wb.close()

"""


