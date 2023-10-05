from openpyxl import load_workbook  #파일 불러오기
wb = load_workbook("sample.xlsx")
ws = wb.active #활성화된 Sheet

# cell 데이터 불러와서 출력하기
for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

# 셀 갯수를 모를 때
for x in range(1, ws.max_row+1):
    for y in range(1, ws.max_column+1):
        print(ws.cell(x,y).value, end=" ")
    print()