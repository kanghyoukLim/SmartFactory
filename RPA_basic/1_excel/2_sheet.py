from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()  # 새로운 sheet 기본이름으로 생성
ws.title = "MySheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB값으로 탭 색 변경

# Sheet, MySheet, YourSheet   인데 위치 지정해서 만들 수 있음
# """
ws1 = wb.create_sheet("YourSheet") #주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet",2) # 2번째 시트에서 만들어짐

new_ws = wb["NewSheet"] # Dic 형태로 Sheet에 접근

print(wb.sheetnames)

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"
# """
wb.save("sample.xlsx")
